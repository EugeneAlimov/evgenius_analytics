from openpyxl import load_workbook

from analytics.models import *


def download_to_base(request):
    """Получаем Excel файл из формы и заполняем базу тегами"""
    ws = ''
    if request.method == 'POST':
        uploaded_file = request.FILES['file']  # получаем содержимое из формы загрузки файлов "name="excel-file""
        wb = load_workbook(filename=uploaded_file, read_only=False)  # получаем excel файл
        ws = wb.active  # делаем лист активным в памяти

    """вычитываем имя группы тегов из каждой строки, пытаемся извлечь из базы имя группы
    соответствующее текущей строчке тега, если имя группы есть - пишем в базу соответствующие теги, привязанные к этой
    группе соотношением один-комногим, если этого имени нет выпадает исключение. Обрабатываем сохраняя
    имя нруппы в таблицу Group и пишем в базу соответствующие теги, привязанные к этой 
    группе соотношением один-комногим
    Если файл с тегами не содержит нужные данные (пустые колонки или ячейки (кроме комментариев)) вывести уведомление 
    об этом"""

    name_group = ''
    groups_list = []

    # for i in range(1, ws.max_row + 1):
    #     if (ws.cell(row=i, column=4)).value.startswith('%'):
    #         groups_list.append(ws.cell(row=i, column=4))
    #     else:
    flag = 0
    for i in range(1, ws.max_row + 1):

        name_tag = (ws.cell(row=i, column=1)).value  # выбираем ячейки из таблицы
        if not isinstance(name_tag, (str, int, bool, float)):
            continue
        address_tag = (ws.cell(row=i, column=4)).value
        name_tag = str(name_tag)
        address_tag = str(address_tag)

        comment_tag = (ws.cell(row=i, column=5)).value
        comment_tag = str(comment_tag)

        if comment_tag == 'None':
            comment_tag = 'No comment'

        """"Для случая когда теги импортированы из таблицы тегов"""

        if address_tag.startswith('%'):
            name_group = (ws.cell(row=i, column=2)).value

            if not Group.objects.filter(name=name_group):
                Group.objects.create(name=name_group)  # сохраняем содержимое в переменной name_group в базу
            name_group_from_db = Group.objects.get(
                name=name_group)  # заносим в переменную объект группы для текущей записи
            if comment_tag.capitalize == 'Spare':
                continue
            tag_table = (ws.cell(row=i, column=2)).value
            data_type = (ws.cell(row=i, column=3)).value
            tag_table = str(tag_table)
            data_type = str(data_type)

            if not Tags.objects.filter(address=address_tag):
                Tags.objects.create(label=name_group_from_db, name_tag=name_tag, tag_table=tag_table,  # сохраняем
                                    data_type=data_type, address=address_tag, comment=comment_tag)  # модель Tags

        else:
            """"Для случая когда теги полуены из DB блоков"""
            if (ws.cell(row=i, column=2)).value == 'Struct':
                if flag == 0:
                    name_group = ''

                if len(name_group) == 0:
                    name_group = (ws.cell(row=i, column=1)).value
                else:
                    name_group = f'{name_group} - {(ws.cell(row=i, column=1)).value}'
                flag += 1
                continue

            if not Group.objects.filter(name=name_group):
                Group.objects.create(name=name_group)  # сохраняем name_group в базу
            print('name_group ', name_group)
            flag = 0

            name_group_from_db = Group.objects.get(
                name=name_group)  # заносим в переменную объект группы для текущей записи
            data_type = (ws.cell(row=i, column=2)).value
            tag_table = (ws.cell(row=i, column=4)).value
            x = ''
            if name_tag == 'InOut' or name_tag == 'Input' or name_tag == 'Output' or name_tag == 'Spare' or name_tag == 'Static' or data_type == 'Struct':
                continue
            if data_type == 'Bool':
                x = '.DBX'
            elif data_type == 'Real' or data_type == 'Time_Of_Day' or data_type == 'Date_And_Time' or data_type == 'DInt' or data_type == 'IEC_TIMER':
                x = '.DBD'
            elif data_type == 'Int' or 'Word':
                x = '.DBW'
            address_tag = f'{(ws.cell(row=i, column=4)).value}{x}{(ws.cell(row=i, column=3)).value}'

            if not Tags.objects.filter(address=address_tag):
                Tags.objects.create(label=name_group_from_db, name_tag=name_tag, tag_table=tag_table,
                                    # сохраняем модель Tags
                                    data_type=data_type, address=address_tag, comment=comment_tag)

    return
